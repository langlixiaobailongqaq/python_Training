#!usr/bin/env python
# -*- coding:utf-8 _*-
"""
@author:zhengxin
@file: day6_11_prcatise1.py
@time: 2022/3/17  14:45
# @describe: 给员⼯⾃动批量发⼯资条
"""
from openpyxl import load_workbook
import smtplib
# 邮件正文
from email.mime.text import MIMEText
# 邮件头
from email.header import Header


# 加载excel文件
wb = load_workbook('小白龙集团-2022年5月工资.xlsx', data_only=True)
sheet = wb.active

# 登录-465/587
# 腾讯企业邮箱为465，此时需要用 smtplib.SMTP_SSL(smtp_host) 连接，
# 其中 smtp_host = ‘smtp.exmail.qq.com’

smtp_obj = smtplib.SMTP('smtp.qq.com', port=587)
# 发件人qq邮箱和授权码
smtp_obj.login('1544717589@qq.com', 'xxxxxx')

# 循环excel
count = 0
# 表头
table_col_html = '<thead>'
for row in sheet.iter_rows(min_row=1):
    count += 1
    # 第一行
    if count == 1:
        for col in row:
            table_col_html += f"<th>{col.value}</th>"
        table_col_html += "<thead/>"
        continue
    else:
        # 开始一行
        row_text = "<tr>"
        for cell in row:
            row_text += f"<td>{cell.value}</td>"
        # 结束一行
        row_text += "</tr>"
        name = row[2]
        staff_email = row[1].value
        print(staff_email, name.value)

    mail_body_context = f"""
                    <h3>{name.value},您好：</h3>
                    <p>请查收您2022-05月的工资条</p>
                    <table border="1px solid black">
                    {table_col_html}
                    {row_text}
                    </table>
                        """
    msg_body = MIMEText(mail_body_context, 'html', 'utf-8')
    # 发送者
    msg_body["From"] = Header("小白龙集团人事部", 'utf-8')
    # 接收者
    msg_body["To"] = Header("小白龙集团员工", 'utf-8')
    # 主题
    msg_body["Subject"] = Header("小白龙集团2022-05月工资", 'utf-8')

    # 发邮件-发件人邮箱、接收人邮箱、邮件标题和内容
    smtp_obj.sendmail('1544717589@qq.com', [staff_email], msg_body.as_string())
    print(f"成功发送工资条到{staff_email}-{name.value}")