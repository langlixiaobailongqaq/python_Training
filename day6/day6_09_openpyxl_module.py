#!usr/bin/env python
# -*- coding:utf-8 _*-
"""
@author:zhengxin
@file: day6_09_openpyxl_module.py
@time: 2022/3/16  17:12
# @describe:  openpyxl 模块
"""
import datetime

import datetime as datetime
from openpyxl import Workbook, load_workbook

""" 新建文件 """
wb = Workbook()
# 获取当前active的sheet
sheet = wb.active
# 打印sheet表名
print(sheet.title)
sheet.title = '葫芦娃'
sheet["B9"] = "black girl"
sheet["C9"] = "171, 48,84 "
sheet['C5'] = '葫芦娃1'
sheet['C7'] = '葫芦娃2'
sheet.append(["Rachel", "170", "49"])
wb.save("excel_test1.xlsx")


# 打开已有文件
wb2 = load_workbook('excel_test.xlsx')


""" 写数据 """
# 方式一：数据可以直接分配到单元格中(可以输入公式)
sheet['C5'] = '葫芦娃1'
sheet['C7'] = '葫芦娃2'

# 方式二：可以附加行，从第一列开始附加(从最下方空白处，最左开始。可以输入多行)
sheet.append([1, 2, 3])

# 方式三：python类型会被自动转换
sheet['A3'] = datetime.datetime.now().strftime('%Y-%m-%d')


""" 选择表 """
# sheet 名称可以作为key 进行索引
wb = load_workbook('excel_test1.xlsx')
wb3 = wb['葫芦娃']
print(wb.sheetnames)
# 获得第1个sheet
sheet = wb.worksheets[0]
print(sheet)

""" 保存表 """
wb.save('文件名称.xlsx')


""" 遍历表数据 """
# 1.按行遍历
# 循环获取表数据
for row in sheet:
    # 循环获取每个单元格数据
    for cell in row:
        print(cell.value, end=',')
    print()

# 2。按列遍历- A1, A2, A3这样的顺序
for column in sheet.columns:
    for cell in column:
        print(cell.value, end=',')
    print()

# 3.遍历指定行数据-从第2⾏开始⾄第5⾏，每⾏打印5列
for row in sheet.iter_rows(min_row=2, max_row=5, max_col=5):
    for cell in row:
        print('行数据:', cell.value, end=',')
    print()

# 4.遍历指定几列的数据-取得第2-第5列的数据
for col in sheet.iter_rows(min_col=2, max_row=5):
    for i in col:
        print('列数据:', i.value, end=',')
    print()


""" 删除工作表 """
# 方式一
# wb.remove(sheet)
# 方式二
# del wb[sheet]


""" 设置单元格样式 """
from openpyxl.styles import Font, colors, Alignment, Border, Side

sheet['B5'] = '字体'
# 字体-
myfont = Font(name='Calibri', size=24, bold=False, italic=False,
              vertAlign=None, underline='none', strike=False, color='00FF0000')
sheet["B5"].font = myfont
sheet["B5"].alignment = Alignment(vertical="center", horizontal="center")

# 第2行行高
sheet.row_dimensions[2].height = 40
# C列列宽
sheet.column_dimensions['C'].width = 30

# 单元格边框border
border = Border(left=Side(border_style='medium', #dark light
                          color='00FF0000'),
                right=Side(border_style='medium',
                          color='00FF0000'),
                top=Side(border_style='medium',
                        color='00FF0000'),
                bottom=Side(border_style='medium',
                            color='00FF0000'),
                diagonal=Side(border_style='medium',
                              color='00FF0000'),
                diagonal_direction=0,
                outline=Side(border_style='medium',
                            color='00FF0000'),
                vertical=Side(border_style='medium',
                              color='00FF0000'),
                horizontal=Side(border_style='medium',
                              color='00FF0000')
              )

sheet["C5"].border = border
wb.save("人数.xlsx")

